import openpyxl
import gspread
from google.oauth2.service_account import Credentials
from collections import defaultdict
import os

# === CONFIGURATION ===
EXCEL_PATH = os.getenv('EXCEL_PATH', '/tmp/BookingData_folder/BookingData.xlsx')
SERVICE_ACCOUNT_FILE = os.getenv('SERVICE_ACCOUNT_FILE', '/tmp/service-account.json')
GSHEET_URL = os.getenv('GSHEET_URL', 'https://docs.google.com/spreadsheets/d/1dp5WINj0Urrvk8Ul2rR_q6HDzjdeAp7iuw5IsY3J3f8/edit#gid=0')
IMP_COMMITMENT_GSHEET_URL = 'https://docs.google.com/spreadsheets/d/1b3VxcaWYkxlBdJlpxefCk4r816eaQl56By2NJlorEQw/edit?gid=667901590#gid=667901590'

# === 1. Read all sheets ===
wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
data_ws = wb['Data']
configs_ws = wb['Configs']

def get_headers(ws):
    return [str(cell.value).strip() for cell in ws[1]]

data_headers = get_headers(data_ws)
configs_headers = get_headers(configs_ws)

# Filter Data for HB/PHB
booking_type_idx = data_headers.index('Booking Type')
data_rows = [
    [cell.value for cell in row]
    for row in data_ws.iter_rows(min_row=2)
    if row[booking_type_idx].value in ('HB', 'PHB')
]

# Forward-fill Package ID/Name in Configs
pkg_id_idx = configs_headers.index('Package ID')
pkg_name_idx = configs_headers.index('Package Name')
configs_rows = []
last_pkg_id = last_pkg_name = None
for row in configs_ws.iter_rows(min_row=2):
    row_values = [cell.value for cell in row]
    if row_values[pkg_id_idx] is not None:
        last_pkg_id = row_values[pkg_id_idx]
    else:
        row_values[pkg_id_idx] = last_pkg_id
    if row_values[pkg_name_idx] is not None:
        last_pkg_name = row_values[pkg_name_idx]
    else:
        row_values[pkg_name_idx] = last_pkg_name
    configs_rows.append(row_values)

# Normalize Package IDs (float->int->str)
def norm_pkgid(val):
    try:
        return str(int(float(val))).strip()
    except:
        return str(val).strip()

def norm_pkgname(val):
    return str(val).strip().lower() if val is not None else ''

# Build lookup for configs by Package ID
configs_lookup = defaultdict(list)
for row in configs_rows:
    pkg_id = norm_pkgid(row[pkg_id_idx])
    configs_lookup[pkg_id].append(row)

# Debug: Print some sample configs data
print(f"✅ Loaded {len(configs_rows)} config rows")
print(f"✅ Found {len(configs_lookup)} unique package IDs in configs")
sample_configs = list(configs_lookup.items())[:3]
for pkg_id, configs in sample_configs:
    print(f"Package ID '{pkg_id}' has {len(configs)} config rows")
    for config in configs[:2]:  # Show first 2 configs per package
        website = config[configs_headers.index('Website')]
        ad_unit = config[configs_headers.index('Ad Unit Type')]
        print(f"  - Website: '{website}', Ad Unit: '{ad_unit}'")

# === 2. Build Sheet2 (expand Data by matching Configs) ===
sheet2_headers = [
    'Expresso ID', 'Campaign Name', 'Package ID', 'Package Name', 'Advertiser',
    'Brand', 'Geo Name', 'Website', 'Section', 'Ad Unit Type', 'Placement'
]
sheet2_rows = []
print(f"✅ Processing {len(data_rows)} data rows")
for i, data_row in enumerate(data_rows):
    pkg_id = norm_pkgid(data_row[data_headers.index('Package ID')])
    if i < 3:  # Debug first 3 data rows
        print(f"Data row {i+1}: Package ID '{pkg_id}'")
    
    if pkg_id in configs_lookup:
        configs_for_pkg = configs_lookup[pkg_id]
        if i < 3:  # Debug first 3 data rows
            print(f"  Found {len(configs_for_pkg)} config rows for this package")
            for j, config in enumerate(configs_for_pkg[:2]):  # Show first 2 configs
                website = config[configs_headers.index('Website')]
                ad_unit = config[configs_headers.index('Ad Unit Type')]
                print(f"    Config {j+1}: Website='{website}', Ad Unit='{ad_unit}'")
        
        for config_row in configs_for_pkg:
            combined_row = [
                data_row[data_headers.index('Expresso ID')],
                data_row[data_headers.index('Campaign Name')],
                pkg_id,
                config_row[pkg_name_idx],
                data_row[data_headers.index('Advertiser')],
                data_row[data_headers.index('Brand')],
                data_row[data_headers.index('Geo Name')],
                config_row[configs_headers.index('Website')],
                config_row[configs_headers.index('Section')],
                config_row[configs_headers.index('Ad Unit Type')],
                config_row[configs_headers.index('Placement')],
            ]
            sheet2_rows.append(combined_row)
    else:
        if i < 3:  # Debug first 3 data rows
            print(f"  ❌ No config found for Package ID '{pkg_id}'")
        # If no config match, fill with blanks for config fields
        combined_row = [
            data_row[data_headers.index('Expresso ID')],
            data_row[data_headers.index('Campaign Name')],
            pkg_id,
            '',
            data_row[data_headers.index('Advertiser')],
            data_row[data_headers.index('Brand')],
            data_row[data_headers.index('Geo Name')],
            '', '', '', ''
        ]
        sheet2_rows.append(combined_row)

print(f"✅ Created {len(sheet2_rows)} Sheet2 rows")

# === 3. Build Final_Innov_Details (parse Website, rename Portal->Publisher, remove TIL_, etc.) ===
def parse_portal_platform(website):
    if not website:
        return '', ''
    
    website_str = str(website).strip()
    
    # Special handling for ET language websites (including AMP variants)
    et_languages = ['Gujarati', 'Hindi', 'Marathi', 'Kannada', 'Bengali', 'Tamil', 'Telugu', 'Malayalam']
    
    # Check if it's an ET language website
    for lang in et_languages:
        if f'ET {lang}' in website_str or f'ET_{lang}' in website_str:
            # Handle AMP case first
            if 'amp' in website_str.lower():
                # For ET language AMP websites, extract the full language name
                if f'ET {lang}' in website_str:
                    portal = f'ET {lang}'
                elif f'ET_{lang}' in website_str:
                    portal = f'ET {lang}'  # Convert underscore to space
                else:
                    # Fallback: extract everything before 'AMP'
                    amp_pos = website_str.lower().find('amp')
                    portal = website_str[:amp_pos].strip()
                return portal, 'Amp'
            # Find the platform identifier and extract the publisher name before it
            platform_identifiers = ['mobile website', 'mobile site', 'android app', 'android apps', 'ios app', 'ios apps', 'mweb', 'website', 'web', 'mobile', 'android', 'aos', 'ios']
            
            # Find which platform identifier is present
            found_platform = ''
            for identifier in platform_identifiers:
                if identifier.lower() in website_str.lower():
                    found_platform = identifier
                    break
            
            # Extract publisher (everything before the platform identifier)
            if found_platform:
                # Find the position of the platform identifier
                platform_pos = website_str.lower().find(found_platform.lower())
                portal = website_str[:platform_pos].strip()
            else:
                # If no platform identifier found, use the whole string
                portal = website_str
            
            # Determine platform based on the found platform identifier
            platform = 'Web'  # Default to 'Web' if no platform identifier found
            if found_platform:
                if any(x in found_platform.lower() for x in ['mobile site', 'mobile website', 'mobile', 'mweb']):
                    platform = 'Mweb'
                elif any(x in found_platform.lower() for x in ['android', 'android app', 'android apps', 'aos']):
                    platform = 'AOS'
                elif any(x in found_platform.lower() for x in ['ios', 'ios app', 'ios apps']):
                    platform = 'IOS'
                elif any(x in found_platform.lower() for x in ['website', 'web']):
                    platform = 'Web'
            return portal, platform
    
    # Handle other AMP websites (non-ET language)
    if 'amp' in website_str.lower():
        # For AMP websites, extract everything before 'AMP'
        amp_pos = website_str.lower().find('amp')
        portal = website_str[:amp_pos].strip()
        return portal, 'Amp'
    
    # Default logic for other websites
    platform_identifiers = ['mobile website', 'mobile site', 'android app', 'android apps', 'ios app', 'ios apps', 'mweb', 'website', 'web', 'mobile', 'android', 'aos', 'ios']
    
    # Find which platform identifier is present
    found_platform = ''
    for identifier in platform_identifiers:
        if identifier.lower() in website_str.lower():
            found_platform = identifier
            break
    
    # Extract publisher (everything before the platform identifier)
    if found_platform:
        # Find the position of the platform identifier
        platform_pos = website_str.lower().find(found_platform.lower())
        portal = website_str[:platform_pos].strip()
    else:
        # If no platform identifier found, use the whole string
        portal = website_str
    
    # Determine platform based on the found platform identifier
    platform = 'Web'  # Default to 'Web' if no platform identifier found
    if found_platform:
        if any(x in found_platform.lower() for x in ['mobile site', 'mobile website', 'mobile', 'mweb']):
            platform = 'Mweb'
        elif any(x in found_platform.lower() for x in ['android', 'android app', 'android apps', 'aos']):
            platform = 'AOS'
        elif any(x in found_platform.lower() for x in ['ios', 'ios app', 'ios apps']):
            platform = 'IOS'
        elif any(x in found_platform.lower() for x in ['website', 'web']):
            platform = 'Web'
    return portal, platform

final_headers = [
    'Expresso ID', 'Campaign Name', 'Package ID', 'Package Name', 'Imp. Commitment',
    'Brand', 'Geo Name', 'Platform', 'Publisher', 'Section', 'Ad Unit Type', 'Placement'
]
final_rows = []
print(f"✅ Processing {len(sheet2_rows)} Sheet2 rows for Final_Innov_Details")

# Debug: Check for any E-TIMES WEBSITE entries
etimes_website_count = 0
etimes_website_rows = []

for i, row in enumerate(sheet2_rows):
    website = row[7]
    ad_unit_type = row[9] or ''
    
    # Check for E-TIMES WEBSITE entries
    if website == 'E-TIMES WEBSITE':
        etimes_website_count += 1
        etimes_website_rows.append((i+1, ad_unit_type))
    
    portal, platform = parse_portal_platform(website)
    
    # Check for Bottom Overlay entries
    if ad_unit_type == 'TIL_Bottom Overlay':
        print(f"🔍 Bottom Overlay Row {i+1}: Website='{website}' → Parsed: Publisher='{portal}', Platform='{platform}'")
    
    if i < 5:  # Debug first 5 rows
        print(f"Row {i+1}: Website='{website}', Ad Unit='{ad_unit_type}'")
        print(f"  Parsed: Publisher='{portal}', Platform='{platform}'")
    
    if ad_unit_type.startswith('TIL_'):
        ad_unit_type = ad_unit_type[4:]
    
    final_rows.append([
        row[0], row[1], row[2], row[3], '', row[5], row[6], platform, portal, row[8], ad_unit_type, row[10]
    ])

print(f"✅ Created {len(final_rows)} Final_Innov_Details rows")

# Report E-TIMES WEBSITE findings
if etimes_website_count > 0:
    print(f"⚠️  Found {etimes_website_count} rows with 'E-TIMES WEBSITE':")
    for row_num, ad_unit in etimes_website_rows:
        print(f"  Row {row_num}: Ad Unit Type = '{ad_unit}'")
else:
    print("✅ No 'E-TIMES WEBSITE' entries found in Sheet2 data")

# === 4. Fetch Impression Commitment from GSheet and merge ===
scopes = ['https://www.googleapis.com/auth/spreadsheets', 'https://www.googleapis.com/auth/drive']
creds = Credentials.from_service_account_file(SERVICE_ACCOUNT_FILE, scopes=scopes)
gc = gspread.authorize(creds)
sh = gc.open_by_url(GSHEET_URL)

def fetch_imp_commitment_data():
    try:
        spreadsheet_id = IMP_COMMITMENT_GSHEET_URL.split('/d/')[1].split('/')[0]
        imp_spreadsheet = gc.open_by_key(spreadsheet_id)
        imp_worksheet = imp_spreadsheet.worksheet('Impression_Commitment')
        imp_data = imp_worksheet.get_all_records()
        print(f"✅ Fetched {len(imp_data)} rows from Impression_Commitment sheet")
        if imp_data:
            print("Available columns in Impression_Commitment sheet:", list(imp_data[0].keys()))
            print("Sample Impression Commitment row:", imp_data[0])
        return imp_data
    except Exception as e:
        print(f"❌ Failed to fetch Impression Commitment data: {e}")
        return []

imp_commitment_data = fetch_imp_commitment_data()

# Explicitly set the expected column names (update these if your sheet uses different names)
IMP_PKGID_COL = 'Til_Package_Id__c'  # or whatever the actual column name is
IMP_GEONAME_COL = 'Geo__c'           # or whatever the actual column name is
IMP_VAL_COL = 'Geo_Level_Imp__c'   # or whatever the actual column name is

imp_lookup = {}
if imp_commitment_data:
    for row in imp_commitment_data:
        pkgid = norm_pkgid(row.get(IMP_PKGID_COL, ''))
        geoname = str(row.get(IMP_GEONAME_COL, '')).strip()
        val = row.get(IMP_VAL_COL, '')
        imp_lookup[(pkgid, geoname)] = val
    print("Sample keys from Impression Commitment lookup:", list(imp_lookup.keys())[:5])

# Add Imp. Commitment to final_rows
sample_main_keys = []
for i, row in enumerate(final_rows):
    key = (norm_pkgid(row[2]), str(row[6]).strip())
    if i < 5:
        sample_main_keys.append(key)
    val = imp_lookup.get(key, '')
    final_rows[i][4] = val
print("Sample keys from main data:", sample_main_keys)

# Show Imp. Commitment only in first row for each Package ID + Geo Name
seen = set()
for i, row in enumerate(final_rows):
    key = (row[2], row[6])
    if key in seen:
        final_rows[i][4] = ''
    else:
        seen.add(key)

# === 5. Upload all sheets to Google Sheets ===
def upload_to_gsheet(rows, headers, sheet_name):
    try:
        try:
            worksheet = sh.worksheet(sheet_name)
            worksheet.clear()
        except gspread.exceptions.WorksheetNotFound:
            worksheet = sh.add_worksheet(title=sheet_name, rows=1000, cols=30)
        worksheet.update([headers] + rows)
        print(f"✅ Uploaded {sheet_name}")
    except Exception as e:
        print(f"❌ Failed to upload {sheet_name}: {e}")

# Prepare Data, Configs, Config2 for upload
# Data
upload_to_gsheet(
    [[cell.value for cell in row] for row in data_ws.iter_rows(min_row=2)],
    data_headers,
    'Data'
)
# Configs
upload_to_gsheet(
    [[cell.value for cell in row] for row in configs_ws.iter_rows(min_row=2)],
    configs_headers,
    'Configs'
)
# Config2
upload_to_gsheet(configs_rows, configs_headers, 'Config2')
# Sheet2
upload_to_gsheet(sheet2_rows, sheet2_headers, 'Sheet2')
# Final_Innov_Details
upload_to_gsheet(final_rows, final_headers, 'Final_Innov_Details')

print("✅ All sheets uploaded to Google Sheets!")