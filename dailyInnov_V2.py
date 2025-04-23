#!/usr/bin/env python3
"""
HB Booking Data Processor for Bitbucket Pipeline
- Processes exported Excel data
- Filters for HB/PHB bookings
- Uploads to specified Google Sheet
"""

import sys
import openpyxl
import gspread
from google.oauth2 import service_account
import os
from datetime import datetime
import json

# Configuration - Get from environment variables
SERVICE_ACCOUNT_JSON = os.getenv('SERVICE_ACCOUNT_JSON')
if not SERVICE_ACCOUNT_JSON:
    raise ValueError("SERVICE_ACCOUNT_JSON environment variable must be set")

GOOGLE_SHEET_URL = os.getenv('GOOGLE_SHEET_URL', 
    'https://docs.google.com/spreadsheets/d/1U12VbADtQ8mQowRjEkEYgxy2bRXDBJwPNKu3OayswIg/edit#gid=1146512691')

SCOPES = [
    'https://www.googleapis.com/auth/spreadsheets',
    'https://www.googleapis.com/auth/drive'
]

def log_message(message, level="INFO"):
    """Standardized logging with timestamps"""
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    print(f"[{timestamp}] [{level}] {message}")

def clean_numeric_values(value):
    """Ensure numeric values are properly formatted"""
    if value is None:
        return None
    try:
        value = float(value)
        # Handle large numbers by limiting decimal places
        return float(f"{value:.2f}") if abs(value) >= 1e6 else value
    except (ValueError, TypeError):
        return value

def load_and_filter_data(excel_path):
    """Load Excel file and filter for HB/PHB bookings"""
    try:
        log_message(f"Loading Excel data from {excel_path}")
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        ws = wb['Data']
        
        # Get headers and their indices
        headers = [cell.value for cell in ws[1]]
        required_columns = {'Expresso ID', 'Campaign Name ', 'Package ID ', 
                          'Package Name ', 'Advertiser ', 'Brand ', 'Value', 'Geo Name ', 'Booking Type'}
        missing_cols = required_columns - set(headers)
        if missing_cols:
            raise ValueError(f"Missing required columns: {missing_cols}")
        
        # Create column index mapping
        col_indices = {header: idx for idx, header in enumerate(headers, 1)}
        
        # Filter for HB/PHB bookings
        hb_data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[col_indices['Booking Type'] - 1] in ['HB', 'PHB']:
                hb_data.append(row)
        
        log_message(f"Found {len(hb_data)} HB & PHB records to process")
        return hb_data, col_indices
        
    except Exception as e:
        log_message(f"Error loading Excel data: {str(e)}", "ERROR")
        raise

def prepare_upload_data(hb_data, col_indices):
    """Prepare data for Google Sheets upload"""
    log_message("Preparing data for upload")
    data_to_upload = []
    for row in hb_data:
        upload_row = [
            str(row[col_indices['Expresso ID'] - 1]),
            str(row[col_indices['Campaign Name '] - 1]).strip(),
            clean_numeric_values(row[col_indices['Package ID '] - 1]),
            str(row[col_indices['Package Name '] - 1]).strip(),
            str(row[col_indices['Advertiser '] - 1]).strip(),
            str(row[col_indices['Brand '] - 1]).strip(),
            clean_numeric_values(row[col_indices['Value'] - 1]),
            str(row[col_indices['Geo Name '] - 1]).strip()
        ]
        data_to_upload.append(upload_row)
    return data_to_upload

def update_google_sheet(data_to_upload):
    """Update Google Sheet with the prepared data"""
    try:
        log_message("Authenticating with Google Sheets")
        
        # Try to get credentials from environment variable first
        try:
            service_account_info = json.loads(SERVICE_ACCOUNT_JSON)
            creds = service_account.Credentials.from_service_account_info(
                service_account_info, scopes=SCOPES)
        except (json.JSONDecodeError, TypeError):
            # If that fails, try to read from file
            service_account_file = '/tmp/service-account.json'
            if not os.path.exists(service_account_file):
                raise ValueError(f"Service account file not found at {service_account_file}")
            creds = service_account.Credentials.from_service_account_file(
                service_account_file, scopes=SCOPES)
        
        gc = gspread.authorize(creds)
        
        log_message(f"Opening Google Sheet: {GOOGLE_SHEET_URL}")
        sheet = gc.open_by_url(GOOGLE_SHEET_URL)
        worksheet = sheet.get_worksheet(0)  # First worksheet
        
        # Find first empty row
        all_values = worksheet.get_all_values()
        first_empty_row = len(all_values) + 1
        
        # Update in batches
        batch_size = 50
        total_rows = len(data_to_upload)
        log_message(f"Uploading {total_rows} rows in batches of {batch_size}")
        
        for i in range(0, total_rows, batch_size):
            batch = data_to_upload[i:i+batch_size]
            range_start = first_empty_row + i
            range_end = range_start + len(batch) - 1
            range_name = f"A{range_start}:H{range_end}"
            
            log_message(f"Uploading rows {range_start}-{range_end}")
            worksheet.update(
                range_name=range_name,
                values=batch
            )
        
        log_message(f"Successfully uploaded {total_rows} HB records starting at row {first_empty_row}")
        return True
        
    except Exception as e:
        log_message(f"Google Sheets error: {str(e)}", "ERROR")
        raise

def main(excel_path):
    """Main processing workflow"""
    try:
        # 1. Load and filter data
        hb_data, col_indices = load_and_filter_data(excel_path)
        
        # 2. Prepare upload data
        data_to_upload = prepare_upload_data(hb_data, col_indices)
        
        # 3. Update Google Sheet
        update_google_sheet(data_to_upload)
        
        log_message("Processing completed successfully")
        return True
        
    except Exception as e:
        log_message(f"Processing failed: {str(e)}", "ERROR")
        return False

if __name__ == "__main__":
    # Get Excel file path from command line or use default
    default_path = os.path.join(os.getenv('DOWNLOAD_DIR', '/tmp/BookingData_folder'), 'BookingData.xlsx')
    excel_path = sys.argv[1] if len(sys.argv) > 1 else default_path
    
    if not os.path.exists(excel_path):
        log_message(f"Excel file not found at {excel_path}", "ERROR")
        sys.exit(1)
    
    if not main(excel_path):
        sys.exit(1)